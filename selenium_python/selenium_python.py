'''
excel练习 写一个自动化测试的代码，步骤如下：
1. 打开百度
2. 输入搜索词
3. 点击搜索
4. 将元素定位和输入的数据使用excel管理
'''

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
from selenium import webdriver
from openpyxl import load_workbook

# wb = Workbook()
#
# dest_filename = 'selenium.xlsx'
#
# ws1 = wb.active
# ws1.title = "range names"
# ws1['A1'] = "元素定位"
# ws1['B1'] = "输入"
# wb.save(filename = dest_filename)

class SeleniumPython:
    def __init__(self):
        ld = load_workbook("./selenium.xlsx")
        self.ws = ld.worksheets[0]



    def go_baidu(self):
        driver = webdriver.Chrome()
        driver.get("https://www.baidu.com/")
        time.sleep(2)
        driver.find_element_by_id(self.ws["A2"].value).send_keys(self.ws["B2"].value)
        driver.find_element_by_id(self.ws["A3"].value).click()
        time.sleeps(15)

if __name__ == '__main__':
    sp = SeleniumPython()
    sp.go_baidu()

